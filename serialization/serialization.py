import pickle
import numpy as np
import time
import matplotlib.pyplot as plt
import openpyxl

# Adjusted function definitions for serialization and deserialization


def serialize_data(data, protocol):
    start_time = time.time()
    buffers = []
    if protocol == 5:
        pickled_data = pickle.dumps(
            data, protocol=protocol, buffer_callback=buffers.append)
    else:
        pickled_data = pickle.dumps(data, protocol=5)
    end_time = time.time()
    # Convert to milliseconds
    return pickled_data, buffers, (end_time - start_time) * 1000


def deserialize_data(pickled_data, protocol, buffers=None):
    start_time = time.time()
    if protocol == 5 and buffers:
        data = pickle.loads(pickled_data, buffers=buffers)
    else:
        data = pickle.loads(pickled_data)
    end_time = time.time()
    return data, (end_time - start_time) * 1000  # Convert to milliseconds

# Adjusted benchmark function


def benchmark(protocol):
    data = np.ones(655360)  # 5MB Object
    pickled_data, buffers, serialize_time = serialize_data(data, protocol)
    _, deserialize_time = deserialize_data(pickled_data, protocol, buffers)
    return serialize_time, deserialize_time


# Perform benchmarks
serialize_time_4, deserialize_time_4 = benchmark(4)
serialize_time_5, deserialize_time_5 = benchmark(5)

# Plotting adjustments
labels = ['In-Band', 'Out-of-Band']
serialize_times = [serialize_time_4, serialize_time_5]
deserialize_times = [deserialize_time_4, deserialize_time_5]

x = np.arange(len(labels))
width = 0.2  # Make the bars narrower

object_size_mb = 5
serialize_throughput_4 = object_size_mb / (serialize_times[0] / 1000)  # MB/s
deserialize_throughput_4 = object_size_mb / \
    (deserialize_times[0] / 1000)  # MB/s
serialize_throughput_5 = object_size_mb / (serialize_times[1] / 1000)  # MB/s
deserialize_throughput_5 = object_size_mb / \
    (deserialize_times[1] / 1000)  # MB/s

throughput_serialize = [serialize_throughput_4, serialize_throughput_5]
throughput_deserialize = [deserialize_throughput_4, deserialize_throughput_5]

# Save throughput values to Excel
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Throughput"

# Write headers
sheet.cell(row=1, column=1, value="Protocol")
sheet.cell(row=1, column=2, value="Serialize Throughput (MB/s)")
sheet.cell(row=1, column=3, value="Deserialize Throughput (MB/s)")

# Write data
for i, label in enumerate(labels, start=2):
    sheet.cell(row=i, column=1, value=label)
    sheet.cell(row=i, column=2, value=throughput_serialize[i-2])
    sheet.cell(row=i, column=3, value=throughput_deserialize[i-2])

# Save the workbook
wb.save("throughput.xlsx")

# # Adjusting the left subplot to display throughput
# fig = plt.figure()
# plt.rcParams.update({'font.size': 14})
# # Throughput Plot for Serialization and Deserialization
# ax1 = fig
# rects1 = ax1.bar(x - width/2, throughput_serialize, width,
#                  label='Serialize', color='skyblue')
# rects2 = ax1.bar(x + width/2, throughput_deserialize, width,
#                  label='Deserialize', color='lightgreen')
#
# ax1.set_ylabel('Throughput (MB/s)')
# ax1.set_title('Throughput of Serializing and Deserializing 5MB Object')
# ax1.set_xticks(x)
# ax1.set_xticklabels(labels)
# ax1.legend(loc='upper left')
# ax1.bar_label(rects1, padding=3, fmt='%.2f')
# ax1.bar_label(rects2, padding=3, fmt='%.2f')
# ax1.set_yscale("log")
# ax1.grid(True, which="both", ls="--", linewidth=0.5)
#
# fig.show()